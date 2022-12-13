import os
from zipfile import ZipFile
import zipfile, glob
import csv
from openpyxl import load_workbook
from PyPDF2 import PdfReader
# archived files
path = (os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources/files'))
expact = (os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources/expact'))
arch_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources/archive1.zip')

with zipfile.ZipFile(arch_path, 'w') as zf:
    for folder, subfolders, files in os.walk(path):
        for file in files:
            s = print(os.path.join(folder, file))
            f = print(os.path.join(folder, file), os.path.relpath(os.path.join(folder, file), path))

            zf.write(os.path.join(folder, file), os.path.relpath(os.path.join(folder, file), path),
                     compress_type=zipfile.ZIP_DEFLATED)


def test_csv():
    with ZipFile(arch_path, 'r') as yzip:
        s = yzip.extract('test2.csv', expact)
        with open(s, newline='') as fp:
            fp = csv.DictReader(fp)
            assert fp.dialect == 'excel'
            fp.__next__()
            assert fp.line_num == 2
        os.remove((os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources/expact/test2.csv')))


def test_xlsx():
    with ZipFile(arch_path, 'r') as xz:
        s = xz.extract('test2.xlsx', expact)
        workbook = load_workbook(s)
        sheet = workbook.active
    assert (sheet.cell(row=3, column=2).value) == 16
    os.remove((os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources/expact/test2.xlsx')))


def test_pdf():
    with ZipFile(arch_path, 'r') as xz:
        s = xz.extract('test3.pdf', expact)
        reader = PdfReader(s)
        page = reader.pages[0]
        text = page.extract_text()
        assert  'Тест пдф файл создала'  in text
